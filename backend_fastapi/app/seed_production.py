from __future__ import annotations

import argparse
import asyncio
import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text

from app.db import SessionLocal
from app.security import hash_password
from app.seed import load_json, parse_date, seed_contacts
from app.seed_reference import sync_attachments_for_session


CENTER_NAMES = {
    "RV": "Rostlinná výroba",
    "ŽV": "Živočišná výroba",
    "ZV": "Živočišná výroba",
    "MECH": "Mechanizace",
    "BPS": "BPS",
    "SS": "Stavební skupina",
    "MINI MLÉKÁRNA": "Mini mlékárna",
    "MINI MLEKARNA": "Mini mlékárna",
    "ŘEDITEL SPOLEČNOSTI": "Ředitelství",
    "REDITEL SPOLECNOSTI": "Ředitelství",
}

WORK_TYPES = [
    "Montáž",
    "Demontáž",
    "Oprava",
    "Údržba",
    "Práce s nakladačem",
    "Krmení",
    "Aplikace hnojiv",
    "Aplikace digestátu",
    "Aplikace POR",
    "Setí",
    "Příprava- před setí",
    "Mulčování",
    "Siláže- řezání",
    "Siláže- odvoz",
    "Nahrnování",
    "Dusání",
    "Sečení",
    "Žně- odvoz",
    "Návoz digestátu",
    "Rozvoz Minimlékarna",
    "Návoz hnojiv",
    "Odkrývání silážních jam",
    "Zakrývání silážních jam",
    "Úklid",
    "Práce BPS",
    "Čištění stroje",
]

SPECIAL_WORK_TYPES = [
    ("Dovolená", "Celodenní nebo půldenní absence z důvodu dovolené"),
    ("Školení", "Účast na školení nebo interní vzdělávání"),
    ("Doktor", "Návštěva lékaře nebo zdravotní volno"),
    ("Darování krve", "Celodenní absence z důvodu darování krve"),
]

LEADER_LEVELS = {"Hlavní vedoucí", "Agronom", "Zootechnička", "Vedoucí střediska", "Vedoucí dílen"}
APPROVED_VIEWER_NAMES = {"Jana Bulíčková", "Jana Bobulová"}


def normalize_center(value: Any) -> str:
    key = str(value or "").strip()
    return CENTER_NAMES.get(key.upper(), key or "Rostlinná výroba")


def role_for_level(level: str) -> str:
    if level == "Schválené výkazy":
        return "approved_viewer"
    if level == "Ředitel":
        return "reditel"
    if level in LEADER_LEVELS:
        return "schvalovatel"
    return "traktorista"


def read_users(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows: list[dict[str, Any]] = []
    current_center = ""
    primary_lead_by_center: dict[str, tuple[str, str]] = {}

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        raw_center, raw_level, raw_name, raw_username, raw_password = row[:5]
        if raw_center:
            current_center = normalize_center(raw_center)
        level = str(raw_level or "").strip()
        full_name = str(raw_name or "").strip()
        username = str(raw_username or "").strip()
        password = str(raw_password or "").strip()
        if not full_name or not username or not password:
            continue

        role = "approved_viewer" if full_name in APPROVED_VIEWER_NAMES else role_for_level(level)
        manager_username = None
        manager_name = None
        if level == "Podřízený" and current_center in primary_lead_by_center:
            manager_username, manager_name = primary_lead_by_center[current_center]
        elif level in LEADER_LEVELS and current_center not in primary_lead_by_center:
            primary_lead_by_center[current_center] = (username, full_name)

        rows.append({
            "id": row_index,
            "username": username,
            "email": f"{username}@lesonice.local",
            "password_hash": hash_password(password),
            "role": role,
            "full_name": full_name,
            "department_name": current_center,
            "scope_department": current_center,
            "position": "",
            "approval_level": level,
            "manager_username": manager_username,
            "manager_name": manager_name,
            "active": True,
        })
    return rows


async def reset_database(session) -> None:
    await session.execute(text("SELECT set_config('adw.actor_name', 'Produkční reset', true)"))
    await session.execute(
        text(
            """
            TRUNCATE
              approvals,
              fuel_entries,
              reports,
              notices,
              machine_service_tasks,
              audit_log,
              users,
              departments,
              work_types,
              fields,
              tractors,
              attachments,
              contacts,
              service_schedule
            RESTART IDENTITY CASCADE
            """
        )
    )


async def seed_departments(session) -> None:
    for code, name in [
        ("REDITELSTVI", "Ředitelství"),
        ("RV", "Rostlinná výroba"),
        ("ZV", "Živočišná výroba"),
        ("MECH", "Mechanizace"),
        ("BPS", "BPS"),
        ("SS", "Stavební skupina"),
        ("MINI_MLEKARNA", "Mini mlékárna"),
    ]:
        await session.execute(
            text(
                """
                INSERT INTO departments(name, code, created_by, updated_by)
                VALUES (:name, :code, 'Produkční seed', 'Produkční seed')
                ON CONFLICT (name) DO UPDATE SET code = EXCLUDED.code, updated_by = 'Produkční seed'
                """
            ),
            {"name": name, "code": code},
        )


async def seed_users(session, users: list[dict[str, Any]]) -> None:
    for user in users:
        await session.execute(
            text(
                """
                INSERT INTO users(
                  id, username, email, password_hash, role, full_name, department_name, scope_department,
                  position, active, manager_username, manager_name, approval_level, created_by, updated_by, last_change
                )
                VALUES (
                  :id, :username, :email, :password_hash, :role, :full_name, :department_name, :scope_department,
                  :position, :active, :manager_username, :manager_name, :approval_level,
                  'Produkční seed', 'Produkční seed', 'Import ostrých účtů'
                )
                ON CONFLICT (username) DO UPDATE SET
                  email = EXCLUDED.email,
                  password_hash = EXCLUDED.password_hash,
                  role = EXCLUDED.role,
                  full_name = EXCLUDED.full_name,
                  department_name = EXCLUDED.department_name,
                  scope_department = EXCLUDED.scope_department,
                  position = EXCLUDED.position,
                  active = EXCLUDED.active,
                  manager_username = EXCLUDED.manager_username,
                  manager_name = EXCLUDED.manager_name,
                  approval_level = EXCLUDED.approval_level,
                  updated_by = 'Produkční seed',
                  last_change = 'Aktualizace ostrého účtu'
                """
            ),
            user,
        )
    await session.execute(text("SELECT setval('users_id_seq', COALESCE((SELECT MAX(id) FROM users), 1), true)"))


async def seed_work_types(session) -> None:
    await session.execute(text("UPDATE work_types SET archived_at = NOW(), archived_by = 'Produkční seed' WHERE archived_at IS NULL"))
    for index, name in enumerate(WORK_TYPES, start=1):
        await session.execute(
            text(
                """
                INSERT INTO work_types(id, name, description, created_by, updated_by, archived_at, archived_by, last_change)
                VALUES (:id, :name, :description, 'Produkční seed', 'Produkční seed', NULL, NULL, 'Import ostrého číselníku')
                ON CONFLICT (name) DO UPDATE SET
                  id = EXCLUDED.id,
                  description = EXCLUDED.description,
                  archived_at = NULL,
                  archived_by = NULL,
                  updated_by = 'Produkční seed',
                  last_change = 'Aktualizace ostrého číselníku'
                """
            ),
            {"id": index, "name": name, "description": name},
        )
    for offset, (name, description) in enumerate(SPECIAL_WORK_TYPES, start=101):
        await session.execute(
            text(
                """
                INSERT INTO work_types(id, name, description, created_by, updated_by, archived_at, archived_by, last_change)
                VALUES (:id, :name, :description, 'Produkční seed', 'Produkční seed', NULL, NULL, 'Interní typ absence')
                ON CONFLICT (name) DO UPDATE SET
                  id = EXCLUDED.id,
                  description = EXCLUDED.description,
                  archived_at = NULL,
                  archived_by = NULL,
                  updated_by = 'Produkční seed',
                  last_change = 'Aktualizace interního typu'
                """
            ),
            {"id": offset, "name": name, "description": description},
        )
    await session.execute(text("SELECT setval('work_types_id_seq', COALESCE((SELECT MAX(id) FROM work_types), 1), true)"))


async def seed_fields(session) -> None:
    for field in load_json("fields.json", []):
        await session.execute(
            text(
                """
                INSERT INTO fields(id, field_code, field_name, quadrant, area, culture, crop, erosion, created_by, updated_by, last_change)
                VALUES (:id, :field_code, :field_name, :quadrant, :area, :culture, :crop, :erosion, 'Produkční seed', 'Produkční seed', 'Import pozemku')
                ON CONFLICT (id) DO UPDATE SET
                  field_code = EXCLUDED.field_code,
                  field_name = EXCLUDED.field_name,
                  quadrant = EXCLUDED.quadrant,
                  area = EXCLUDED.area,
                  culture = EXCLUDED.culture,
                  crop = EXCLUDED.crop,
                  erosion = EXCLUDED.erosion,
                  updated_by = 'Produkční seed'
                """
            ),
            {
                "id": field.get("id"),
                "field_code": field.get("field_code") or str(field.get("id")),
                "field_name": field.get("field_name") or field.get("name") or f"Pozemek {field.get('id')}",
                "quadrant": field.get("quadrant"),
                "area": field.get("area"),
                "culture": field.get("culture"),
                "crop": field.get("crop"),
                "erosion": field.get("erosion"),
            },
        )
    await session.execute(text("SELECT setval('fields_id_seq', COALESCE((SELECT MAX(id) FROM fields), 1), true)"))


async def seed_tractors(session) -> None:
    for tractor in load_json("tractors.json", []):
        await session.execute(
            text(
                """
                INSERT INTO tractors(id, tractor_code, tractor_name, service_centers, vehicle_type, status, created_by, updated_by, last_change)
                VALUES (:id, :tractor_code, :tractor_name, :service_centers, :vehicle_type, :status, 'Produkční seed', 'Produkční seed', 'Import stroje')
                ON CONFLICT (id) DO UPDATE SET
                  tractor_code = EXCLUDED.tractor_code,
                  tractor_name = EXCLUDED.tractor_name,
                  service_centers = EXCLUDED.service_centers,
                  vehicle_type = EXCLUDED.vehicle_type,
                  status = EXCLUDED.status,
                  updated_by = 'Produkční seed'
                """
            ),
            {
                "id": tractor.get("id"),
                "tractor_code": tractor.get("tractor_code") or tractor.get("code") or str(tractor.get("id")),
                "tractor_name": tractor.get("tractor_name") or tractor.get("name") or f"Stroj {tractor.get('id')}",
                "service_centers": tractor.get("service_centers") or [],
                "vehicle_type": tractor.get("vehicle_type") or "traktor",
                "status": tractor.get("status") or "active",
            },
        )
    await session.execute(text("SELECT setval('tractors_id_seq', COALESCE((SELECT MAX(id) FROM tractors), 1), true)"))


async def seed_service_schedule(session) -> None:
    for service in load_json("service-schedule.json", []):
        await session.execute(
            text(
                """
                INSERT INTO service_schedule(date, workshop, bps_service, bps_feeding)
                VALUES (:date, CAST(:workshop AS jsonb), CAST(:bps_service AS jsonb), CAST(:bps_feeding AS jsonb))
                ON CONFLICT (date) DO UPDATE SET
                  workshop = EXCLUDED.workshop,
                  bps_service = EXCLUDED.bps_service,
                  bps_feeding = EXCLUDED.bps_feeding
                """
            ),
            {
                "date": parse_date(service.get("date")),
                "workshop": json.dumps(service.get("workshop")),
                "bps_service": json.dumps(service.get("bps_service")),
                "bps_feeding": json.dumps(service.get("bps_feeding")),
            },
        )


async def seed_production(login_xlsx: Path, reset: bool) -> None:
    users = read_users(login_xlsx)
    if not users:
        raise RuntimeError("Přihlašovací XLSX neobsahuje žádné uživatele.")

    async with SessionLocal() as session:
        await session.execute(text("SELECT set_config('adw.actor_name', 'Produkční seed', true)"))
        if reset:
            await reset_database(session)
        await seed_departments(session)
        await seed_users(session, users)
        await seed_work_types(session)
        await seed_fields(session)
        await seed_tractors(session)
        await sync_attachments_for_session(session)
        await seed_contacts(session)
        await seed_service_schedule(session)
        await session.commit()
    print(f"Production seed complete: {len(users)} users, {len(WORK_TYPES)} work types.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", nargs="?", default=os.environ.get("ADW_LOGIN_XLSX", "/app/runtime/Přihlašovací udaje.xlsx"))
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()
    asyncio.run(seed_production(Path(args.xlsx), args.reset))


if __name__ == "__main__":
    main()
